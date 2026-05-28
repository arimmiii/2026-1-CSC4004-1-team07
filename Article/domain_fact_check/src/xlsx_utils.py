from __future__ import annotations

import zipfile
import xml.etree.ElementTree as ET
from pathlib import Path
from typing import Any
import re

try:
    from openpyxl import load_workbook  # type: ignore
except ImportError:  # pragma: no cover
    load_workbook = None


def load_xlsx_rows(path: str, sheet_name: str = "", limit: int = 0) -> list[dict[str, Any]]:
    if load_workbook is not None:
        return _load_with_openpyxl(path=path, sheet_name=sheet_name, limit=limit)
    return _load_with_zip_xml(path=path, sheet_name=sheet_name, limit=limit)


def _load_with_openpyxl(path: str, sheet_name: str = "", limit: int = 0) -> list[dict[str, Any]]:
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb[sheet_name] if sheet_name else wb[wb.sheetnames[0]]
    rows = ws.iter_rows(values_only=True)
    headers = [str(cell).strip() if cell is not None else "" for cell in next(rows)]
    items: list[dict[str, Any]] = []
    for index, row in enumerate(rows, 1):
        item = {headers[i]: row[i] for i in range(min(len(headers), len(row)))}
        items.append(item)
        if limit and index >= limit:
            break
    return items


def _load_with_zip_xml(path: str, sheet_name: str = "", limit: int = 0) -> list[dict[str, Any]]:
    ns = {
        "a": "http://schemas.openxmlformats.org/spreadsheetml/2006/main",
        "r": "http://schemas.openxmlformats.org/officeDocument/2006/relationships",
        "pr": "http://schemas.openxmlformats.org/package/2006/relationships",
    }
    p = Path(path)
    with zipfile.ZipFile(p) as z:
        shared: list[str] = []
        if "xl/sharedStrings.xml" in z.namelist():
            root = ET.fromstring(z.read("xl/sharedStrings.xml"))
            for si in root.findall("a:si", ns):
                shared.append("".join(t.text or "" for t in si.findall(".//a:t", ns)))

        wb = ET.fromstring(z.read("xl/workbook.xml"))
        rels = ET.fromstring(z.read("xl/_rels/workbook.xml.rels"))
        rel_map = {rel.attrib["Id"]: rel.attrib["Target"] for rel in rels.findall("pr:Relationship", ns)}

        target = ""
        for s in wb.find("a:sheets", ns):
            name = s.attrib["name"]
            if sheet_name and name != sheet_name:
                continue
            rid = s.attrib["{http://schemas.openxmlformats.org/officeDocument/2006/relationships}id"]
            target = "xl/" + rel_map[rid]
            break
        if not target:
            raise ValueError(f"Sheet not found: {sheet_name}")

        root = ET.fromstring(z.read(target))
        rows = root.findall(".//a:sheetData/a:row", ns)
        decoded_rows: list[list[Any]] = []
        for row in rows:
            values_by_index: dict[int, Any] = {}
            max_index = -1
            for cell in row.findall("a:c", ns):
                ref = cell.attrib.get("r", "")
                col_index = _excel_col_to_index(ref)
                max_index = max(max_index, col_index)
                cell_type = cell.attrib.get("t")
                value_node = cell.find("a:v", ns)
                value: Any = ""
                if cell_type == "inlineStr":
                    inline = cell.find("a:is", ns)
                    if inline is not None:
                        value = "".join(t.text or "" for t in inline.findall(".//a:t", ns))
                elif value_node is not None:
                    value = value_node.text or ""
                    if cell_type == "s":
                        value = shared[int(value)] if value.isdigit() and int(value) < len(shared) else value
                values_by_index[col_index] = value
            values = [values_by_index.get(i, None) for i in range(max_index + 1)] if max_index >= 0 else []
            decoded_rows.append(values)
            if limit and len(decoded_rows) - 1 >= limit:
                break

    if not decoded_rows:
        return []
    headers = [str(cell).strip() if cell is not None else "" for cell in decoded_rows[0]]
    items: list[dict[str, Any]] = []
    for row in decoded_rows[1:]:
        items.append({headers[i]: row[i] if i < len(row) else None for i in range(len(headers))})
    return items


def _excel_col_to_index(ref: str) -> int:
    match = re.match(r"([A-Z]+)", ref)
    if not match:
        return 0
    letters = match.group(1)
    index = 0
    for ch in letters:
        index = index * 26 + (ord(ch) - ord("A") + 1)
    return index - 1
